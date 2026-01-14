# pls_project/io_utils.py
import re
import pandas as pd
import numpy as np
from scipy.stats import norm
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

LIKERT_MAP = {
    "非常不同意": 1, "很不同意": 1,
    "不同意": 2,
    "普通": 3, "中立": 3,
    "同意": 4,
    "非常同意": 5, "很同意": 5,
}

_FULLWIDTH = str.maketrans("０１２３４５６７８９．－＋", "0123456789.-+")

def _norm_text(s: str) -> str:
    return str(s).translate(_FULLWIDTH).replace("\u3000", " ").strip()

def to_score(x):
    if pd.isna(x):
        return np.nan
    if isinstance(x, (int, float, np.integer, np.floating)):
        return float(x)

    s = _norm_text(x)

    if s in LIKERT_MAP:
        return float(LIKERT_MAP[s])

    # 抓字串中第一個數字：例如 "5（非常同意）" / "同意(4)"
    m = re.search(r"[-+]?\d+(\.\d+)?", s)
    if m:
        try:
            return float(m.group(0))
        except Exception:
            return np.nan

    return np.nan

def reverse_1to5(x):
    if pd.isna(x):
        return np.nan
    try:
        v = float(x)
    except Exception:
        return np.nan
    if v < 1 or v > 5:
        return np.nan
    return 6 - v

def safe_t_p(est: np.ndarray, se: np.ndarray):
    est = np.asarray(est, dtype=float)
    se  = np.asarray(se, dtype=float)
    t = np.divide(np.abs(est), se, out=np.full_like(est, np.nan), where=(se > 0))
    p = np.where(np.isfinite(t), 2 * (1 - norm.cdf(t)), np.nan)
    return t, p

def get_or_create_ws(writer, sheet_name: str):
    if sheet_name in writer.book.sheetnames:
        ws = writer.book[sheet_name]
    else:
        ws = writer.book.create_sheet(sheet_name)
    writer.sheets[sheet_name] = ws
    return ws

def _autofit_columns(ws, start_row: int, start_col: int, nrows: int, ncols: int, max_width: int = 50):
    for j in range(start_col, start_col + ncols):
        letter = get_column_letter(j)
        best = 10
        for i in range(start_row, start_row + nrows):
            v = ws.cell(row=i, column=j).value
            if v is None:
                continue
            best = max(best, min(max_width, len(str(v)) + 2))
        ws.column_dimensions[letter].width = best

def write_block(writer, sheet_name, ws, startrow, title, df_block, index=False, autofit=True):
    if df_block is None or (isinstance(df_block, pd.DataFrame) and df_block.empty):
        df_block = pd.DataFrame([{"Info": "No data"}])

    ws.cell(row=startrow + 1, column=1, value=title).font = Font(bold=True, size=12)

    df_block = df_block.copy()
    df_block.to_excel(writer, sheet_name=sheet_name, index=index, startrow=startrow + 1, startcol=0)

    header_row = startrow + 2
    ncols = df_block.shape[1] + (1 if index else 0)
    for c in range(1, ncols + 1):
        ws.cell(row=header_row, column=c).font = Font(bold=True)

    if autofit:
        nrows = 1 + len(df_block) + 1
        _autofit_columns(ws, start_row=startrow + 1, start_col=1, nrows=nrows + 2, ncols=ncols)

    return startrow + 1 + 1 + len(df_block) + 2
